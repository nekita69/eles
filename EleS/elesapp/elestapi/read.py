import os
import openpyxl as op


'''
#Получаем все файлы которые есть в директории; (Тут только ведомости)
docks_path = 'docks/' # Для приложения - 'docks/'
files = os.listdir(docks_path) #Все файлы, которые нахоядся в директории 'elesapp/docks';
for file in files:
    res = file.split('.')
    if res[len(res)-1] == 'xlsx': #Проверка, то что файл EXCEL;
        listbooks.append(op.open(docks_path + file)) #Открываем файлы и помещаем в список;
'''

def open_docks():
    listbooks = [] #Список открытых файлов;
    #Получаем все файлы которые есть в директории; (Тут только ведомости)
    docks_path = 'docks/' # Для приложения - 'docks/'
    files = os.listdir(docks_path)
    for file in files:
        res = file.split('.')
        if res[len(res)-1] == 'xlsx': #Проверка, то что файл EXCEL;
            listbooks.append(op.open(docks_path + file)) #Открываем файлы и помещаем в список;
    return listbooks #Список открытых файлов;


def close_docks(listbooks): #Закрываем 
    docks_path = 'docks/' # Для приложения - 'docks/'
    files = os.listdir(docks_path)
    for i in range(len(listbooks)):
        listbooks[i].save(docks_path + files[i])
    return True #Закрыли все файлы;



# Метод получает ФИО учителя и возвращает предметы которые он ведет
def get_subject(teacher_name):
    listbooks = open_docks() #Список открытых файлов;
    
    listsub = [] #Список для результата;
    for book in listbooks: #Идем по открытым документам EXCEL;
        for i in range (len(book.worksheets)): #Идем по листам документа EXCEL
            sheet = book.worksheets[i]
            step = 4
            temp = []
            while(sheet.cell(row = 11, column = step).value != None): #Пока на листе не закончатся ФИО;
                if (sheet.cell(row = 11, column = step).value == teacher_name):
                    flag =  True
                    temprow = 14
                    while(sheet.cell(row = temprow, column = 2).value != None): #Пока есть студенты
                        if(sheet.cell(row = temprow, column = step).value == None or sheet.cell(row = temprow, column = step).value == 'п/а'):
                            temprow += 1
                            continue
                        else:
                            flag = False
                            break

                    if(flag):
                        listsub.append({
                            'm_group':sheet.cell(row = 3, column = 3).value,
                            'cource':book.worksheets[i].title,
                            "subject":sheet.cell(row = 9, column = step).value, 
                            "group":sheet.cell(row = 5, column = 3).value,
                            "mark":sheet.cell(row = 13, column = step).value
                            })
                step += 1

    close_docks(listbooks) #Закрываем все excel файлы;
    return listsub #Список json с дисциплиной, оценкой, группой, направлением




def get_group_stud(m_gr, gr, dischip = "", mark = ""):
    listbooks = open_docks()
    #направление: m_gr = 00.00.00 first second
    #Метод для получения списка группы;
    docks_path = 'docks/' # Для приложения - 'docks/'
    files = os.listdir(docks_path)
    
    tmp_gr = gr.strip(" ")
    for i in range(len(files)): #Имем по файлам в директории
        for j in range(len(listbooks[i].worksheets)): #Идем по листам документа
            sheet = listbooks[i].worksheets[j]
            t_gr = sheet.cell(row=5, column=3).value.strip(" ")
            if(t_gr == tmp_gr): #Если нашли нужную группу
                cols_mark = 4 #Колонка с оценками
                while(sheet.cell(row=9, column=cols_mark).value != None):
                    if(sheet.cell(row=9, column=cols_mark).value == dischip and sheet.cell(row=13, column=cols_mark).value == mark):
                        break
                    cols_mark += 1
                
                
                studs = [] #Список для результата
                nn = 14
                fio = sheet.cell(row=nn, column=2).value
                number = sheet.cell(row=nn, column=1).value
                mark = sheet.cell(row=nn, column=1).value
                while fio != None:
                    if(sheet.cell(row=nn, column=cols_mark).value != 'п/а'):
                        studs.append({'name':fio, 'number':number})

                    nn += 1
                    fio = sheet.cell(row=nn, column=2).value
                    number = sheet.cell(row=nn, column=1).value
                close_docks(listbooks)
                return studs
    close_docks(listbooks)
    return None



def fill_file(gr, dischip, mark, group_list): #Группа, дисциплина, оценка, список группы
    listbooks = open_docks()
    
    for book in listbooks:
        gr = gr.strip(" ")
        for sheet in book.worksheets:
            gr_list = sheet.cell(row=5, column=3).value.strip(" ")
            if(gr_list == gr): #Находим группу;
                col = 4
                while(sheet.cell(row=9, column=col).value != None):
                    #Находим совпадение дисциплины и оценку
                    dischip = dischip.strip(" ")
                    mark = mark.strip(" ")
                    if(sheet.cell(row=9, column=col).value == dischip and sheet.cell(row=13, column=col).value == mark):
                        mrow = 14
                        item = 0
                        while(item < len(group_list)): #Пока не закончится список
                            name = group_list[item]['name']
                            if(sheet.cell(row=mrow, column=2).value == name):
                                sheet.cell(row=mrow, column=col).value = group_list[item]['mark']
                                mrow += 1
                                item += 1
                            else:
                                mrow += 1
                    col += 1
                    
    close_docks(listbooks) #Закрываем файлы


    #Выставление в зачетку
    docks_path1 = 'elzach/' # Для приложения - 'docks/'
    files1 = os.listdir(docks_path1)

    new_mark = ''
    if(mark == 'Зачёт с оценкой'):
        new_mark = 'з/о'
    elif(mark == 'Зачёт'):
        new_mark = 'з'
    elif(mark == 'Экзамен'):
        new_mark = 'экз'
    else:
        new_mark = 'к/р'

    for fileitem in files1:
        fle = fileitem.split(".")[0]
        print(fle, " ", gr, " ", gr == fle)
        if fle == gr: #Нашли файл с зачетками группы;
            listbook1 = op.open(docks_path1 + fileitem) #Открываем файл
            for sheet in listbook1.worksheets:
                step = 0 #По списку группы;
                item = 1 #По строчкам таблицы;
                while(sheet.cell(row=item, column=2).value != 'учится'): #Находим первую непустую строчку
                    item += 1
                #print(item)

                name = group_list[step]['name'].split(" ")
                #print(name)
                while sheet.cell(row=item, column=4).value != None:
                    dis = sheet.cell(row=item, column=9).value.split(" (")[0] #dis
                    if(sheet.cell(row=item, column=6).value.strip(" ") == name[0] and
                       sheet.cell(row=item, column=7).value.strip(" ") == name[1] and
                       sheet.cell(row=item, column=8).value.strip(" ") == name[2] and
                       dis == dischip):
                        #Выставить оценку, перейти на другое имя
                        
                        mr = group_list[step]['mark']
                        if(mr == 'зачет'):
                            mr = 'зачтено'
                        elif(mr == '5'):
                            mr = 'отлично'
                        elif(mr == '4'):
                            mr = 'хорошо'
                        elif(mr == '3'):
                            mr = 'удовлетворительно'
                        else:
                            mr = 'долг'


                        sheet.cell(row=item, column=12).value = mr
                        #Сделать оценку
                        step += 1
                        name = group_list[step]['name'].split(" ")
                        #print(name)
                    item += 1
            #print(item)
            #print(step)
            listbook1.save(docks_path1 + fileitem) #Сохраняем файл
            return True
    return False
                    
                   
















   






# Метод получает Фио учителя и предмет который он ведет и возвращает номер группы
def get_group(teacher_name, subject):
    for book in (listbooks):
        listint2 = []
        for i in range (len(book.worksheets)):
            sheet = book.worksheets[i]
            step = 4
            while True:
                if (sheet.cell(row = 11, column = step).value == teacher_name):
                    if (sheet.cell(row = 9, column = step).value == subject):
                        inf = {'name': teacher_name, 'group': sheet.cell(5,3).value}
                        listint2.append(inf)
                        break
                if (sheet.cell(row = 11, column = step).value == None):
                    break
                step += 1
        # for i in range (len(listint2)):
        #     print(listint2[i])

#get_group(teach_name,'Иностранный язык')


# Метод получает номер группы и колонку преподователя и возвращает список студентов группы
def get_students(group, col):
    for book in listbooks:
        liststud = []
        for i in range (len(book.worksheets)):
            temp = book.worksheets[i].title
            if (temp.find(group) != 1):
                sheet = book[temp]
                break
        roww = 14
        while True:
            s_inf = {'s_name': sheet.cell(roww, 2).value, 'n_z': sheet.cell(roww,3).value, 'grade': sheet.cell(roww, col).value}
            liststud.append(s_inf)
            roww += 1
            if (sheet.cell(roww, 2).value == None):
                break
        # for i in range (len(liststud)):
        #     print(liststud[i])

#get_students('ПИ-С-22', 4)



